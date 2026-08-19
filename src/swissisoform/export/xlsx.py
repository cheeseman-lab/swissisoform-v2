"""Export a paired-TIS parquet to a collaborator-friendly Excel workbook.

Produces a .xlsx with three sheets:
  - ``isoforms``        — one row per TIS, identity columns first.
  - ``data_dictionary`` — every column with its group + a plain-English
                          description, generated from the column-naming
                          convention (scope_module_field).
  - ``key_columns``     — a curated high-signal subset for a quick read.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Glossary ───────────────────────────────────────────────────────────────

SCOPE = {
    "canonical": "Canonical protein",
    "isoform": "Isoform protein",
    "cmp": "Comparison (isoform vs canonical)",
    "diff": "Differential region",
    "expr": "Expression",
}

MODULE = {
    "biophysics": "biophysics",
    "motifs": "linear sequence motifs",
    "localization": "DeepLoc 2 subcellular localization",
    "clinical": "clinical variants (gnomAD / ClinVar / COSMIC)",
    "signalp": "SignalP 6 signal-peptide prediction",
    "targetp": "TargetP 2 N-terminal targeting prediction",
    "interproscan": "InterProScan 6 domains / families",
    "massspec": "mass-spec support (tryptic peptides + PepQuery)",
    "conservation": "Zoonomia conservation (PhyloP / PhastCons + primate/mammal frame integrity)",
    "core": "core TIS identity metadata",
    "initiation": "translation-initiation (Kozak) context",
    "plm": "protein language model variant effect (ESM-C masked-marginal LLR)",
    "structure": "Boltz-2 predicted structure metrics",
    "variant": "genomic variant ↔ ORF-region intersection",
    "varianteffect": "per-variant effect on identified mutants (ESM-C ΔLLR + AlphaMissense)",
    "scoring": "dual-axis evidence score (existence C+D / functional L+M+P+S)",
    "len": "protein length (aa)",
    "transcript": "transcript identifier",
}

# Specific field humanizations (matched as a substring of the field tail).
FIELD = {
    "pI": "isoelectric point (pI)",
    "gravy": "GRAVY hydropathy index",
    "disorder": "predicted disorder fraction",
    "llps_score": "liquid–liquid phase-separation propensity",
    "instability_index": "protein instability index",
    "aromaticity": "aromaticity",
    "fraction_charged": "fraction of charged residues",
    "fraction_disorder_promoting": "fraction of disorder-promoting residues",
    "prionlike_fraction": "prion-like sequence fraction",
    "shannon_entropy": "sequence Shannon entropy",
    "normalized_complexity": "normalized sequence complexity",
    "longest_homopolymer": "longest homopolymer run",
    "aa_diversity": "distinct amino-acid types in the N-terminal 20 residues",
    "pipi_propensity": "pi–pi contact propensity",
    "rg_fg_density": "Arg/Gly–Phe/Gly (RG/FG) density",
    "frac_intact": "fraction of species with an intact reading frame",
    "n_species_aligned": "number of species aligned over the region",
    "n_species_intact_frame": "number of species with an intact reading frame",
    "start_codon_conserved": "start codon conserved across species",
    "mean_pident": "mean amino-acid percent identity vs human",
    "deepest_species": "deepest-diverging species still frame-intact",
    "max_depth": "phylogenetic depth of the deepest frame-intact species",
    "phylop": "PhyloP conservation score (per-base, 241-mammal)",
    "phastcons": "PhastCons conserved-element probability",
    "plddt_isoform_mean": "mean Boltz pLDDT confidence over the isoform",
    "plddt_canonical_mean": "mean Boltz pLDDT over the canonical",
    "plddt_diffregion_mean": "mean Boltz pLDDT over the differential region",
    "plddt_diffregion_std": "pLDDT std-dev over the differential region",
    "plddt_delta_shared": "Δ mean pLDDT over the shared region (isoform − canonical)",
    "tm_score": "TM-score: isoform-vs-canonical fold similarity (1 = identical)",
    "rmsd_global": "tm-align global RMSD over the canonical-vs-isoform alignment "
    "(Å; not strictly shared-residue-only)",
    "extension_contacts": "inter-residue contacts made by the extension",
    "deeploc_prediction": "predicted subcellular compartment",
    "deeploc_signals": "predicted sorting signals",
    "deeploc_membrane": "predicted membrane association",
    "total_variants": "total variants overlapping the protein",
    "llr": "masked-marginal log-likelihood ratio (variant effect)",
    "raw_count": "raw Ribo-TISH read count",
    "cpm": "counts per million (RNA-normalized)",
    "p_value": "Ribo-TISH detection p-value",
    "initiation_efficiency": "translation-initiation efficiency",
    # generic / structural fields
    "hits": "raw per-hit list (positions + metadata); truncated in Excel, full in the parquet",
    "summary": "aggregate summary (counts / status)",
    "status": "module run status (ok / not_run / no_alignment / …)",
    "backend": "structure-prediction backend (boltz / chai)",
    # signalp / targetp
    "sp_prob": "signal-peptide probability",
    "mtp_prob": "mitochondrial transit-peptide probability",
    "ctp_prob": "chloroplast transit-peptide probability",
    "cleavage_site": "predicted cleavage-site position",
    "prediction": "predicted class / label",
    "probability": "prediction probability",
    # core identity
    "in_frame": "True if the isoform ORF is in-frame with the canonical",
    "large_truncation_warning": "flag: unusually large N-terminal truncation",
    "orf_type": "ORF type relative to canonical",
    "canonical_protein_length": "canonical protein length (aa)",
    "isoform_protein_length": "isoform protein length (aa)",
    "differential_length_aa": "differential-region length (aa)",
    "variant_id": "TIS identifier",
    # initiation context
    "kozak_context": "13-nt Kozak context (start codon at indices 9–11)",
    "kozak_hamming_major": "Hamming distance to the strong-Kozak consensus",
    "kozak_hamming_partial": "Hamming distance to a partial-Kozak match",
    "kozak_hamming_full": "Hamming distance to the full-Kozak consensus",
    "kozak_window_gc_content": "GC content of the 13-nt Kozak window around the TIS",
    # variant intersection
    "n_total": "total intersecting variants",
    "n_in_unique_region": "variants within the isoform-unique region",
    "n_in_shared_region": "variants within the shared region",
    "n_pathogenic_in_unique_region": "pathogenic variants in the unique region",
    "n_pathogenic_in_shared_region": "pathogenic variants in the shared region",
    # plm vep
    "constraint_delta": (
        "mean logP(wt) unique − shared; positive = unique region better predicted, "
        "i.e. more conserved"
    ),
    "n_constrained_positions_unique": "conserved positions in the unique region",
    "n_constrained_positions_shared": "conserved positions in the shared region",
    # variant effect (per-variant ESM-C ΔLLR + AlphaMissense)
    "n_scored_plm": "clinical variants given an ESM-C ΔLLR score",
    "n_scored_am": "clinical variants given an AlphaMissense score",
    "n_scorable_in_unique": "scorable (missense, predictor-available) variants "
    "in the unique region",
    "n_damaging_in_unique": "damaging variants in the unique region "
    "(AlphaMissense likely_pathogenic or ESM-C ΔLLR ≤ −7.5)",
    "mean_delta_llr_unique": "mean ESM-C ΔLLR (logP(alt)−logP(wt)) over unique-region variants",
    "min_delta_llr_unique": "most damaging ESM-C ΔLLR over unique-region variants",
    "mean_am_pathogenicity_unique": "mean AlphaMissense pathogenicity over unique-region variants",
    "max_am_pathogenicity_unique": "max AlphaMissense pathogenicity over unique-region variants",
    "n_am_pathogenic_in_unique": "AlphaMissense likely_pathogenic variants in the unique region",
    # conservation BigWig (distinct per region so they don't collapse to one desc)
    "phylop_at_tis": "PhyloP at the start codon",
    "phylop_kozak_mean": "mean PhyloP over the Kozak window",
    "phylop_unique_region_mean": "mean PhyloP over the isoform-unique region",
    "phylop_shared_region_mean": "mean PhyloP over the shared region",
    "phylop_enrichment": "PhyloP unique/shared enrichment",
    "phastcons_at_tis": "PhastCons at the start codon",
    "phastcons_kozak_mean": "mean PhastCons over the Kozak window",
    "phastcons_unique_region_mean": "mean PhastCons over the isoform-unique region",
    "phastcons_shared_region_mean": "mean PhastCons over the shared region",
    "phastcons_enrichment": "PhastCons unique/shared enrichment",
}

# Exact descriptions for identity / metadata columns.
EXACT = {
    "gene_name": "HGNC gene symbol",
    "gene_id": "Ensembl gene ID",
    "tis_id": "TIS identifier — chrom:position:strand:start_codon:transcript",
    "transcript_id": "Ensembl transcript the isoform is defined on",
    "chrom": "chromosome",
    "position": "genomic position of the start codon (0-based)",
    "strand": "genomic strand (+/−)",
    "start_codon": "start codon — ATG or a near-cognate (CTG/GTG/ACG/…)",
    "orf_type": "ORF type vs canonical (Extended/Truncated/uORF/Internal/Novel/Annotated)",
    "aa_len": "isoform protein length (amino acids)",
    "differential_sequence": "amino-acid sequence of the isoform-unique (differential) region",
    "diff_region_confidence": "confidence in the differential-region assignment",
    "diff_start": "differential-region start (residue index in the diff_space protein)",
    "diff_end": "differential-region end (residue index in the diff_space protein)",
    "diff_space": "protein coordinate space for diff_start/end: 'isoform' (extensions/uORFs) "
    "or 'canonical' (truncations, where the unique region is the lost N-terminus)",
    "kozak_context": "13-nt Kozak context around the start (ATG/near-cognate at indices 9–11)",
    "tis_pvalue": "Ribo-TISH TIS-calling p-value",
    "ribo_pvalue": "Ribo-TISH ribosome-signal p-value",
    "fisher_qvalue": "Fisher combined q-value (multiple-testing adjusted)",
}

SUFFIX = [
    ("_delta", " — Δ (isoform − canonical)"),
    ("_unique", " — over the isoform-unique region"),
    ("_shared", " — over the shared region"),
    ("_ratio", " — unique/shared ratio"),
    ("_enriched", " — True if enriched in the unique region"),
    ("_changed", " — True if the call differs between canonical and isoform"),
    ("_canonical", " — canonical value"),
    ("_isoform", " — isoform value"),
]


def _humanize(text: str) -> str:
    return text.replace("_", " ").strip()


_MAX_CELL = 2000  # Excel's hard cap is 32767; keep cells readable.


def _trim_cell(v: object) -> object:
    """Stringify list/array cells; truncate anything too long for a clean Excel cell."""
    if isinstance(v, str):
        s = v
    elif isinstance(v, (list, tuple)) or type(v).__name__ == "ndarray":
        s = str(list(v))
    else:
        return v
    if len(s) > _MAX_CELL:
        return s[:_MAX_CELL] + f" …[+{len(s) - _MAX_CELL} chars — full value in the parquet]"
    return s


def describe(col: str) -> tuple[str, str]:
    """Return (group, description) for a column name."""
    if col.startswith("generef_"):
        field = col[len("generef_") :]
        gdesc = {
            "uniprot_id": "UniProt accession (reviewed, human)",
            "function": "Affinage mechanistic narrative (PMID-cited)",
            "subcellular_location": "Affinage subcellular location(s)",
            "keywords": "Affinage functional terms (molecular activity + pathway)",
        }.get(field, _humanize(field))
        return "gene reference", f"Gene-level reference — {gdesc}"
    if col in EXACT:
        grp = "identity" if not col.startswith("diff_") else "differential region"
        return grp, EXACT[col]

    # expr_<cellline>_<field>
    if col.startswith("expr_"):
        rest = col[len("expr_") :]
        for fld in ("raw_count", "cpm", "p_value", "initiation_efficiency"):
            if rest.endswith(fld):
                cell = rest[: -len(fld) - 1]
                return "expression", f"{FIELD.get(fld, _humanize(fld))} in {cell}"
        return "expression", _humanize(rest)

    scope = None
    rest = col
    for pre in ("canonical", "isoform", "cmp", "diff"):
        if col.startswith(pre + "_"):
            scope, rest = pre, col[len(pre) + 1 :]
            break
    if scope is None:
        return "metadata", _humanize(col)

    module = rest.split("_", 1)[0]
    field = rest[len(module) + 1 :] if "_" in rest else ""
    mod_label = MODULE.get(module, module)
    scope_label = SCOPE.get(scope, scope)

    # Strip a leading token that just repeats the module or a known sub-namespace
    # (signalp_signalp_prob, core_identity_orf_type, plm_vep_status, …).
    for red in (module, "identity", "context", "vep", "intersection", "frame"):
        if field.startswith(red + "_"):
            field = field[len(red) + 1 :]
            break
    clade = ""
    if field.startswith("primate_"):
        clade, field = "primates — ", field[len("primate_") :]
    elif field.startswith("mammalian_"):
        clade, field = "mammals — ", field[len("mammalian_") :]

    # field description: longest matching key in FIELD, else humanized
    field_desc = None
    for key in sorted(FIELD, key=len, reverse=True):
        if key in field:
            field_desc = FIELD[key]
            break
    if field_desc is None:
        field_desc = _humanize(field) or mod_label

    # suffix note — skip if the description already conveys that region/sense
    suffix_note = ""
    for suf, note in SUFFIX:
        s = suf.lstrip("_")
        if field.endswith(s):
            if s in ("unique", "shared") and s in field_desc.lower():
                break
            suffix_note = note
            break

    desc = f"{scope_label} — {mod_label}: {clade}{field_desc}{suffix_note}"
    return f"{scope}/{module}", desc


# ── Column ordering ──────────────────────────────────────────────────────

IDENTITY_ORDER = [
    "gene_name", "gene_id", "tis_id", "transcript_id", "chrom", "position",
    "strand", "start_codon", "orf_type", "aa_len", "kozak_context",
    "tis_pvalue", "ribo_pvalue", "fisher_qvalue",
    "diff_start", "diff_end", "diff_space",
    "diff_region_confidence", "differential_sequence",
]

KEY_COLUMNS = [
    "gene_name", "tis_id", "transcript_id", "orf_type", "strand", "start_codon",
    "aa_len", "kozak_context",
    "isoform_scoring_existence_score", "isoform_scoring_existence_high_confidence",
    "isoform_scoring_functional_score", "isoform_scoring_functional_high_confidence",
    "isoform_localization_deeploc_prediction",
    "cmp_localization_deeploc_prediction_changed",
    "cmp_clinical_n_hits_in_diff_region",
    "isoform_varianteffect_n_scorable_in_unique",
    "isoform_varianteffect_n_damaging_in_unique",
    "isoform_varianteffect_mean_delta_llr_unique",
    "isoform_varianteffect_n_am_pathogenic_in_unique",
    "isoform_conservation_frame_primate_frac_intact",
    "isoform_conservation_frame_mammalian_frac_intact",
    "isoform_structure_plddt_isoform_mean", "isoform_structure_tm_score",
    "cmp_biophysics_pI_delta", "cmp_biophysics_gravy_delta",
]


def order_columns(cols: list[str]) -> list[str]:
    """Order columns: identity first, then expression, scoring, everything else."""
    rank = {}
    for c in cols:
        if c in IDENTITY_ORDER:
            rank[c] = (0, IDENTITY_ORDER.index(c))
        elif c.startswith("expr_"):
            rank[c] = (1, c)
        elif c.startswith("isoform_scoring"):
            rank[c] = (2, c)
        else:
            rank[c] = (3, c)
    return sorted(cols, key=lambda c: rank[c])


# ── Workbook writing ───────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, n_cols: int) -> None:
    for j in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=j)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{ws.max_row}"


def write_workbook(df: pd.DataFrame, out_path: Path) -> None:
    """Write the three-sheet workbook (isoforms, data_dictionary, key_columns)."""
    cols = order_columns(list(df.columns))
    data = df[cols].apply(lambda col: col.map(_trim_cell))

    dict_rows = []
    for c in cols:
        g, d = describe(c)
        dict_rows.append({"column": c, "group": g, "description": d})
    ddf = pd.DataFrame(dict_rows)

    key_cols = [c for c in KEY_COLUMNS if c in df.columns]
    kdf = df[key_cols].apply(lambda col: col.map(_trim_cell))

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        data.to_excel(xw, sheet_name="isoforms", index=False)
        ddf.to_excel(xw, sheet_name="data_dictionary", index=False)
        kdf.to_excel(xw, sheet_name="key_columns", index=False)

        wb = xw.book
        _style_header(wb["isoforms"], len(cols))
        _style_header(wb["data_dictionary"], 3)
        _style_header(wb["key_columns"], len(key_cols))
        # widen dictionary columns for readability
        dd = wb["data_dictionary"]
        dd.column_dimensions["A"].width = 48
        dd.column_dimensions["B"].width = 22
        dd.column_dimensions["C"].width = 90
        for c in dd.iter_rows(min_row=2, max_col=3):
            c[2].alignment = Alignment(wrap_text=True, vertical="top")
        wb["key_columns"].column_dimensions["A"].width = 14
        wb["key_columns"].column_dimensions["B"].width = 42
